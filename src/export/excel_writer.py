import os
from datetime import datetime
from openpyxl import Workbook


class ExcelExporter:
    def __init__(self):
        self.default_output_dir = os.path.join(os.getcwd(), 'output')
        if not os.path.exists(self.default_output_dir):
            os.makedirs(self.default_output_dir)

    def generate_filename(self, base_name):
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        name_without_ext = os.path.splitext(base_name)[0]
        return f"{name_without_ext}_{timestamp}.xlsx"

    def export(self, df, filename=None, sheet_name='Data'):
        if filename is None:
            filename = self.generate_filename('output')
        
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        output_path = os.path.join(self.default_output_dir, filename)
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            for r_idx, row in enumerate(df.itertuples(index=False), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            for col_idx, col_name in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
            
            wb.save(output_path)
            return output_path
        except Exception as e:
            raise Exception(f"Error al exportar a Excel: {str(e)}")

    def get_output_dir(self):
        return self.default_output_dir
